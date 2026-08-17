#!/usr/bin/env python3
"""
5_verify_fields.py  —  Re-read flagged cells and update with verified values.

This is the VERIFICATION step. After initial extraction, many cells are
flagged uncertain ([?], [??], [faded], [illegible]). This script:

    1. Lists all flagged cells for a given PDF.
    2. Accepts a JSON file of corrections (from a re-read at higher resolution).
    3. Applies the updates, coloring each cell green [verified], or keeping
       its flag if still uncertain.
    4. Appends a verification note to the Progress sheet.

Usage:
    # Show what's flagged
    python 5_verify_fields.py --pdf "OPD BOOK NOVEMBER 2025.pdf" --list

    # Apply corrections
    python 5_verify_fields.py --pdf "OPD BOOK NOVEMBER 2025.pdf" \\
           --updates corrections.json --through-page 7

corrections.json format — array of [page, record_no, {field: value}]:
    [
      [1, "78", {"Patient name": "Tuori Tolwa [verified]"}],
      [1, "79", {"Patient name": "Maiga Pakwana [verified]",
                 "Village": "Mweya (code5) [verified]"}],
      [2, "84", {"Patient name": "Nankinga Praise [verified]"}]
    ]

Values should include markers:
    [verified]   — confirmed correct   → cell turns green
    [?]          — still uncertain      → stays yellow
    [??]         — very uncertain       → stays orange
    (no marker)  — assumed high conf    → no fill
"""

import sys, os, json, argparse

from extraction_helpers import (
    master_workbook, update_record_fields,
    flagged_cells_for_pdf, update_progress,
)


def list_flagged(pdf):
    """Print all flagged cells for a PDF."""
    from config import MASTER_XLSX
    import openpyxl
    wb = openpyxl.load_workbook(MASTER_XLSX, data_only=False)
    cells = flagged_cells_for_pdf(wb, pdf)
    wb.close()

    if not cells:
        print(f"No flagged cells for '{pdf}'.")
        return

    print(f"Flagged cells for '{pdf}' ({len(cells)} total):\n")
    cur_page = None
    for c in sorted(cells, key=lambda x: (x["page"] or 0, x["record_no"] or 0)):
        if c["page"] != cur_page:
            cur_page = c["page"]
            print(f"  ── Page {cur_page} ──")
        print(f"    Rec {c['record_no']:>4}  {c['field']:<20}  "
              f"= {c['value']}  ({c['color']})")


def apply_updates(pdf, updates_path, through_page=None):
    """Load corrections JSON and apply to the master workbook."""
    with open(updates_path) as f:
        updates = json.load(f)

    with master_workbook() as wb:
        total = 0
        no_match = []
        for entry in updates:
            page, rec_no, fields = entry[0], entry[1], entry[2]
            n = update_record_fields(wb, pdf, page, rec_no, fields)
            if n == 0:
                no_match.append((page, rec_no))
            total += n

        # Append progress note
        if "Progress" in wb.sheetnames:
            ws = wb["Progress"]
            for row in ws.iter_rows(min_row=2):
                if str(row[0].value) == pdf:
                    cur = str(row[5].value or "")
                    add = f"V:through p{through_page}" if through_page else "V:partial"
                    row[5].value = (cur + " | " + add).strip(" |")
                    break

    print(f"Updated {total} rows for '{pdf}'.")
    if no_match:
        print(f"WARNING: {len(no_match)} entries not matched:")
        for pg, rn in no_match:
            print(f"  page={pg} rec={rn}")


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Verify and update flagged cells.")
    p.add_argument("--pdf", required=True)
    p.add_argument("--list", action="store_true",
                   help="Just list flagged cells, don't update")
    p.add_argument("--updates", default=None,
                   help="Path to corrections JSON")
    p.add_argument("--through-page", type=int, default=None,
                   help="Last page verified (for progress note)")
    args = p.parse_args()

    if args.list:
        list_flagged(args.pdf)
    elif args.updates:
        if not os.path.isfile(args.updates):
            sys.exit(f"ERROR: {args.updates} not found")
        apply_updates(args.pdf, args.updates, args.through_page)
    else:
        print("Specify --list or --updates <file>. See --help.")
