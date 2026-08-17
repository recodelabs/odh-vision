#!/usr/bin/env python3
"""
8_progress_report.py  —  Report overall pipeline progress.

Usage:
    python 8_progress_report.py           → compact: 4 percentages
    python 8_progress_report.py -v        → verbose breakdown
"""

import sys, os, csv, argparse
from config import MASTER_XLSX, AUDIT_LOG, TOTAL_PDFS
import openpyxl


def read_progress(wb):
    counts = {"extracted": 0, "optimized": 0, "verified": 0,
              "total": TOTAL_PDFS, "pdfs": []}
    if "Progress" not in wb.sheetnames:
        return counts
    ws = wb["Progress"]
    for r in range(2, ws.max_row + 1):
        pdf = ws.cell(r, 1).value
        if not pdf:
            continue
        ext  = str(ws.cell(r, 2).value or "").strip().lower()
        opt  = str(ws.cell(r, 3).value or "").strip().lower()
        recs = ws.cell(r, 4).value
        notes = str(ws.cell(r, 6).value or "")
        if ext == "yes":   counts["extracted"] += 1
        if opt == "yes":   counts["optimized"] += 1
        if "verified-done" in notes.lower(): counts["verified"] += 1
        counts["pdfs"].append({"pdf": pdf, "extracted": ext,
                                "optimized": opt, "records": recs,
                                "notes": notes})
    return counts


def read_audit_accuracy():
    if not os.path.exists(AUDIT_LOG) or os.path.getsize(AUDIT_LOG) == 0:
        return {"match": 0, "mismatch": 0, "accuracy": 0}
    m = mm = 0
    with open(AUDIT_LOG, newline="") as f:
        for row in csv.DictReader(f):
            if row["result"] == "MATCH":    m += 1
            elif row["result"] == "MISMATCH": mm += 1
    judged = m + mm
    return {"match": m, "mismatch": mm,
            "accuracy": round(m / judged * 100, 1) if judged else 0}


def report(verbose=False):
    wb = openpyxl.load_workbook(MASTER_XLSX, data_only=True)
    prog  = read_progress(wb)
    wb.close()
    audit = read_audit_accuracy()

    t = prog["total"]
    ext = round(prog["extracted"] / t * 100, 1)
    opt = round(prog["optimized"] / t * 100, 1)
    ver = round(prog["verified"]  / t * 100, 1)
    aud = audit["accuracy"]

    # Compact format (S's preference)
    print(f"{ext}%, {opt}%, {ver}%, {aud}%")

    if verbose:
        print(f"\n  Extraction:    {prog['extracted']}/{t} = {ext}%")
        print(f"  Optimization:  {prog['optimized']}/{t} = {opt}%")
        print(f"  Verification:  {prog['verified']}/{t}  = {ver}%")
        print(f"  Audit accuracy: {audit['match']}/({audit['match']}+{audit['mismatch']}) = {aud}%")

        incomplete = [p for p in prog["pdfs"] if p["extracted"] != "yes"]
        if incomplete:
            print(f"\n  Not yet extracted ({len(incomplete)}):")
            for p in incomplete[:20]:
                print(f"    {p['pdf']}  [{p['extracted']}]  "
                      f"recs={p['records'] or 0}")
            if len(incomplete) > 20:
                print(f"    ... and {len(incomplete) - 20} more")


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Pipeline progress report.")
    p.add_argument("-v", "--verbose", action="store_true")
    args = p.parse_args()
    report(verbose=args.verbose)
