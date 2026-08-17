"""
extraction_helpers.py  —  Shared library for the ODH extraction pipeline.

All the reusable functions live here so the numbered scripts stay short.

Sections:
    1. File locking          — acquire_lock()
    2. Workbook management   — master_workbook(), region_sheets(), etc.
    3. Record CRUD           — append_records(), update_record_fields(), update_progress()
    4. Discovery             — next_pdf_to_extract(), pdfs_with_flags()
    5. Cell inspection       — cell_rgb(), is_flagged(), flagged_cells_for_pdf()
    6. Image helpers         — enhance_region()
    7. Legend                — update_legend()
"""

import os, shutil, time, contextlib, datetime, re, fcntl
from typing import List, Dict, Optional, Tuple

# Import config FIRST — it auto-installs missing packages
from config import (
    MASTER_XLSX, MASTER_BAK, MASTER_LOCK,
    HEADERS, HEADER_INDEX, AUDIT_FIELDS,
    FLAG_COLORS, VERIFIED_COLOR, RESERVED_SHEETS,
    TOTAL_PDFS, OPTIMIZED_DIR, CONFIDENCE_PALETTE,
)

import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image, ImageEnhance, ImageFilter


# ═══════════════════════════════════════════════════════════════════════════════
#  1. File locking  (advisory — prevents two processes writing at once)
# ═══════════════════════════════════════════════════════════════════════════════

@contextlib.contextmanager
def acquire_lock(timeout: float = 30.0):
    """
    Advisory file lock around the master workbook.

    Usage:
        with acquire_lock():
            wb = openpyxl.load_workbook(MASTER_XLSX)
            ...
            wb.save(MASTER_XLSX)

    Raises TimeoutError if the lock isn't acquired within *timeout* seconds.
    """
    deadline = time.time() + timeout
    fd = os.open(MASTER_LOCK, os.O_CREAT | os.O_RDWR)
    try:
        while True:
            try:
                fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
                break
            except BlockingIOError:
                if time.time() > deadline:
                    raise TimeoutError(
                        f"Could not acquire lock on {MASTER_LOCK} "
                        f"within {timeout}s")
                time.sleep(0.5)
        yield
    finally:
        fcntl.flock(fd, fcntl.LOCK_UN)
        os.close(fd)


# ═══════════════════════════════════════════════════════════════════════════════
#  2. Workbook management
# ═══════════════════════════════════════════════════════════════════════════════

@contextlib.contextmanager
def master_workbook(read_only: bool = False):
    """
    Context manager: load the master xlsx, yield the workbook, then
    (unless read_only) save + back up.

    Uses advisory locking so concurrent scheduled runs don't clobber
    each other.

    Usage:
        with master_workbook() as wb:
            append_records(wb, rows)
    """
    if read_only:
        wb = openpyxl.load_workbook(MASTER_XLSX, data_only=True)
        try:
            yield wb
        finally:
            wb.close()
        return

    with acquire_lock():
        wb = openpyxl.load_workbook(MASTER_XLSX, data_only=False)
        try:
            yield wb
        finally:
            # Atomic save: write to .tmp, then replace
            tmp = MASTER_XLSX + ".wtmp"
            wb.save(tmp)
            os.replace(tmp, MASTER_XLSX)
            shutil.copy2(MASTER_XLSX, MASTER_BAK)
            wb.close()


def region_sheets(wb) -> list:
    """Return worksheets that contain patient data (skip Legend / Progress)."""
    return [wb[n] for n in wb.sheetnames if n not in RESERVED_SHEETS]


def _ensure_region_sheet(wb, region_name: str):
    """Get or create a data sheet for *region_name* with the header row."""
    if region_name in wb.sheetnames:
        return wb[region_name]
    ws = wb.create_sheet(region_name)
    for ci, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=ci, value=h)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
#  3. Record CRUD
# ═══════════════════════════════════════════════════════════════════════════════

def append_records(wb, rows: List[Dict]) -> int:
    """
    Append record dicts to the correct region sheet(s).

    Each dict must have "Region".  Keys matching HEADERS are written;
    unknowns are silently ignored.  Returns the number of rows appended.
    """
    count = 0
    for rec in rows:
        region = rec.get("Region", "Unknown")
        ws = _ensure_region_sheet(wb, region)
        next_row = ws.max_row + 1
        for header, ci in HEADER_INDEX.items():
            val = rec.get(header)
            if val is not None:
                ws.cell(row=next_row, column=ci + 1, value=val)
        count += 1
    return count


def update_record_fields(wb, pdf: str, page, record_no,
                         updates: Dict[str, str]) -> int:
    """
    Update specific fields on an existing record identified by
    (Source PDF, Page, Record No.).

    *updates* maps field name → new value (may include markers like
    "[verified]", "[?]").  Applies the matching confidence fill color
    automatically.

    Returns the number of rows matched (0 or 1).
    """
    for ws in region_sheets(wb):
        for row in ws.iter_rows(min_row=2):
            if len(row) <= 4:
                continue
            r_pdf = str(row[HEADER_INDEX["Source PDF"]].value or "")
            r_page = row[HEADER_INDEX["Page"]].value
            r_rec = str(row[HEADER_INDEX["Record No."]].value or "")

            if r_pdf != pdf:
                continue
            if str(r_page) != str(page):
                continue
            if r_rec != str(record_no):
                continue

            # Found the row — apply updates
            for field, value in updates.items():
                ci = HEADER_INDEX.get(field)
                if ci is None:
                    continue
                cell = row[ci]
                cell.value = value
                # Apply color based on marker
                cell.fill = _fill_for_marker(value)
            return 1
    return 0


def _fill_for_marker(value: str) -> PatternFill:
    """Choose a cell fill color based on confidence markers in the value."""
    v = str(value)
    if "[verified]" in v:
        return GREEN_FILL
    elif "[illegible]" in v:
        return RED_FILL
    elif "[??]" in v:
        return ORANGE_FILL
    elif "[faded]" in v:
        return BLUE_FILL
    elif "[?]" in v:
        return YELLOW_FILL
    else:
        return PatternFill()   # no fill = high confidence


def update_progress(wb, source_pdf: str, extracted: str = "",
                    optimized: str = "", records_count: int = 0,
                    notes: str = ""):
    """
    Create or update the Progress-sheet row for *source_pdf*.

    Columns:  A=Source PDF  B=Extracted  C=Optimized  D=Records  E=Verified  F=Notes
    """
    if "Progress" not in wb.sheetnames:
        ws = wb.create_sheet("Progress")
        for ci, h in enumerate(
                ["Source PDF", "Extracted", "Optimized",
                 "Records", "Verified", "Notes"], start=1):
            ws.cell(row=1, column=ci, value=h)
    else:
        ws = wb["Progress"]

    target_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == source_pdf:
            target_row = r
            break
    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(target_row, 1, source_pdf)

    if extracted:
        ws.cell(target_row, 2, extracted)
    if optimized:
        ws.cell(target_row, 3, optimized)
    if records_count:
        ws.cell(target_row, 4, records_count)
    if notes:
        ws.cell(target_row, 6, notes)


# ═══════════════════════════════════════════════════════════════════════════════
#  4. Discovery
# ═══════════════════════════════════════════════════════════════════════════════

def next_pdf_to_extract(wb) -> Optional[str]:
    """Return the first PDF filename whose Extracted column is not 'yes',
    or None if everything is done."""
    if "Progress" not in wb.sheetnames:
        return None
    ws = wb["Progress"]
    for r in range(2, ws.max_row + 1):
        pdf = ws.cell(r, 1).value
        status = str(ws.cell(r, 2).value or "").strip().lower()
        if pdf and status != "yes":
            return pdf
    return None


def pdfs_with_flags(wb) -> List[str]:
    """Return a list of PDF filenames that still have flagged (uncertain) cells."""
    flagged = set()
    for ws in region_sheets(wb):
        for row in ws.iter_rows(min_row=2):
            if len(row) <= 3:
                continue
            pdf = row[HEADER_INDEX["Source PDF"]].value
            if not pdf or pdf in flagged:
                continue
            for ci in range(min(len(row), len(HEADERS))):
                rgb = cell_rgb(row[ci])
                if rgb and rgb in FLAG_COLORS:
                    flagged.add(pdf)
                    break
    return sorted(flagged)


def find_pdf_path(pdf_filename: str) -> Optional[str]:
    """
    Walk the ODHFILESCANS_optimized/ tree and return the full path to
    *pdf_filename*, or None if not found.
    """
    for root, dirs, files in os.walk(OPTIMIZED_DIR):
        for f in files:
            if f == pdf_filename:
                return os.path.join(root, f)
    return None


# ═══════════════════════════════════════════════════════════════════════════════
#  5. Cell inspection
# ═══════════════════════════════════════════════════════════════════════════════

def cell_rgb(cell) -> Optional[str]:
    """Return the ARGB hex string of a cell's fill, or None."""
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.rgb:
        rgb = str(fill.fgColor.rgb)
        if rgb != "00000000":
            return rgb
    return None


def is_flagged(cell) -> bool:
    """True if this cell's fill is a FLAG_COLOR (uncertain / problem)."""
    rgb = cell_rgb(cell)
    return rgb is not None and rgb in FLAG_COLORS


def flagged_cells_for_pdf(wb, pdf_name: str) -> List[Dict]:
    """
    Return every flagged cell for records of *pdf_name*.

    Each dict: {sheet, row, page, record_no, field, value, color}
    """
    results = []
    for ws in region_sheets(wb):
        for row in ws.iter_rows(min_row=2):
            if len(row) <= 3 or str(row[2].value) != pdf_name:
                continue
            for ci in range(min(len(row), len(HEADERS))):
                rgb = cell_rgb(row[ci])
                if rgb and rgb in FLAG_COLORS:
                    results.append({
                        "sheet": ws.title,
                        "row": row[ci].row,
                        "page": row[3].value,
                        "record_no": row[0].value,
                        "field": HEADERS[ci],
                        "value": row[ci].value,
                        "color": rgb,
                    })
    return results


def confidence_stats(wb, pdf_name: str) -> Tuple[int, int]:
    """Return (confident_cells, total_cells) for *pdf_name*."""
    conf = total = 0
    for ws in region_sheets(wb):
        for row in ws.iter_rows(min_row=2):
            if len(row) <= 3 or str(row[2].value) != pdf_name:
                continue
            for ci in range(min(len(row), len(HEADERS))):
                total += 1
                if not is_flagged(row[ci]):
                    conf += 1
    return conf, total


# ═══════════════════════════════════════════════════════════════════════════════
#  6. Image helpers
# ═══════════════════════════════════════════════════════════════════════════════

def enhance_region(src_path: str, out_path: str,
                   box: Tuple[int, int, int, int] = None,
                   scale: float = 2.0,
                   mode: str = "contrast") -> str:
    """
    Crop, upscale, and enhance a region for better readability.

    box   : (left, top, right, bottom) pixel coords, or None = full image
    scale : upscale factor (2.0 = double resolution)
    mode  : "contrast"  — sharpen + contrast boost (best for handwriting)
            "threshold" — convert to binary B&W
    """
    im = Image.open(src_path)
    if box:
        im = im.crop(box)
    if scale != 1.0:
        im = im.resize(
            (int(im.width * scale), int(im.height * scale)),
            Image.LANCZOS)

    if mode == "contrast":
        im = ImageEnhance.Contrast(im).enhance(1.8)
        im = ImageEnhance.Sharpness(im).enhance(2.0)
    elif mode == "threshold":
        im = im.convert("L")
        im = im.point(lambda p: 255 if p > 140 else 0, "1")

    im.save(out_path)
    return out_path


# ═══════════════════════════════════════════════════════════════════════════════
#  7. Legend
# ═══════════════════════════════════════════════════════════════════════════════

# Pre-built PatternFill objects — importable by other scripts
GREEN_FILL  = PatternFill(start_color=VERIFIED_COLOR, end_color=VERIFIED_COLOR, fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
BLUE_FILL   = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

_FILL_MAP = {
    "Green":  GREEN_FILL,
    "Yellow": YELLOW_FILL,
    "Blue":   BLUE_FILL,
    "Orange": ORANGE_FILL,
    "Red":    RED_FILL,
}


def update_legend(wb):
    """
    Write (or refresh) the color-confidence legend on the Legend sheet.
    Safe to call repeatedly — it overwrites the palette section.
    """
    if "Legend" not in wb.sheetnames:
        wb.create_sheet("Legend")
    ws = wb["Legend"]

    start = ws.max_row + 2
    ws.cell(row=start, column=1,
            value="Color palette (graduated confidence)")
    for k, (name, marker, desc) in enumerate(CONFIDENCE_PALETTE, start=start + 1):
        c = ws.cell(row=k, column=1, value=name)
        c.fill = _FILL_MAP[name]
        ws.cell(row=k, column=2, value=f"{marker} — {desc}")

    note_row = start + len(CONFIDENCE_PALETTE) + 2
    ws.cell(row=note_row, column=1, value="Verification pass")
    ws.cell(row=note_row, column=2,
            value="Cells flagged uncertain are re-read from a full-resolution "
                  "crop with image enhancement. Confirmed reads turn green; "
                  "genuinely unclear cells keep their flag — never guessed.")
